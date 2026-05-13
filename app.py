import os
import io
import json
import base64
import logging
import zipfile
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for)
from werkzeug.security import check_password_hash, generate_password_hash
import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production")

# ── Audit logger ───────────────────────────────────────────────────────────────
LOG_PATH = Path(__file__).parent / "logs" / "audit.log"
LOG_PATH.parent.mkdir(exist_ok=True)

audit_logger = logging.getLogger("audit")
audit_logger.setLevel(logging.INFO)
audit_logger.propagate = False

_fh = logging.FileHandler(LOG_PATH, encoding="utf-8")
_fh.setFormatter(logging.Formatter("%(asctime)s\t%(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
audit_logger.addHandler(_fh)

def audit(action: str, detail: str = ""):
    user = session.get("username", "—")
    ip   = request.remote_addr or "—"
    audit_logger.info(f"{user}\t{ip}\t{action}\t{detail}")


# ── Users ──────────────────────────────────────────────────────────────────────
def load_users() -> dict:
    """Словарь {username: password_hash} из users.json рядом с app.py."""
    users_path = Path(__file__).parent / "users.json"
    if users_path.exists():
        with open(users_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


# ── Auth decorator ─────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


# ── PostgreSQL ─────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host":     os.environ.get("DB_HOST",     "localhost"),
    "port":     int(os.environ.get("DB_PORT", 5432)),
    "dbname":   os.environ.get("DB_NAME",     "your_database"),
    "user":     os.environ.get("DB_USER",     "your_user"),
    "password": os.environ.get("DB_PASSWORD", "your_password"),
}

SQL_DIR = Path(__file__).parent / "sql_scripts"


def load_scripts():
    registry_path = Path(__file__).parent / "scripts" / "registry.json"
    if registry_path.exists():
        with open(registry_path, encoding="utf-8") as f:
            return json.load(f)
    scripts = []
    for sql_file in sorted(SQL_DIR.glob("*.sql")):
        scripts.append({
            "id":          sql_file.stem,
            "name":        sql_file.stem.replace("_", " ").title(),
            "description": "",
            "file":        sql_file.name,
        })
    return scripts


def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)


def run_sql_script(script_file: str, date_from: str, date_to: str,
                   use_time: bool = True, time_from: str = "00:00:00", time_to: str = "23:59:59"):
    sql_path = SQL_DIR / script_file
    if not sql_path.exists():
        raise FileNotFoundError(f"Script not found: {script_file}")

    sql = sql_path.read_text(encoding="utf-8")

    d_from = datetime.strptime(date_from, "%Y-%m-%d")
    d_to   = datetime.strptime(date_to,   "%Y-%m-%d")

    if use_time:
        fmt_from = d_from.strftime("%d.%m.%Y") + f" {time_from}"
        fmt_to   = d_to.strftime("%d.%m.%Y")   + f" {time_to}"
    else:
        fmt_from = d_from.strftime("%d.%m.%Y")
        fmt_to   = d_to.strftime("%d.%m.%Y")

    sql = sql.replace(":date_from", f"'{fmt_from}'")
    sql = sql.replace(":date_to",   f"'{fmt_to}'"  )

    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql)
            rows    = cur.fetchall()
            columns = [desc[0] for desc in cur.description] if cur.description else []
            return columns, [dict(r) for r in rows]
    finally:
        conn.close()


def _safe_val(row, col_name):
    val = row.get(col_name)
    if hasattr(val, "tzinfo") and val.tzinfo is not None:
        val = val.replace(tzinfo=None)
    return val


def write_sheet(ws, columns: list, rows: list,
                styled: bool = False, title: str = "", date_from: str = "", date_to: str = ""):
    """Write SQL output. styled=True adds header title, column formatting and alternating rows."""

    if styled:
        # ── Style constants ────────────────────────────────────────────────────
        HEADER_FILL  = PatternFill("solid", start_color="1E3A5F")
        HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ALT_FILL     = PatternFill("solid", start_color="EEF2F7")
        BORDER_SIDE  = Side(style="thin", color="C8D0DC")
        CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                              top=BORDER_SIDE,  bottom=BORDER_SIDE)
        BODY_FONT    = Font(name="Calibri", size=10)
        BODY_ALIGN   = Alignment(vertical="center", wrap_text=True)
        HDR_ALIGN    = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Title row
        ws.append([f"{title}   |   {date_from} — {date_to}"])
        title_cell = ws.cell(1, 1)
        title_cell.font      = Font(name="Calibri", bold=True, size=13, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.append([])  # blank spacer

        # Column header row
        header_row = 3
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(header_row, col_idx, col_name)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.border = CELL_BORDER; cell.alignment = HDR_ALIGN
        ws.row_dimensions[header_row].height = 22

        # Data rows
        for row_idx, row in enumerate(rows, start=header_row + 1):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row_idx, col_idx, _safe_val(row, col_name))
                cell.font = BODY_FONT; cell.border = CELL_BORDER; cell.alignment = BODY_ALIGN
                if fill:
                    cell.fill = fill

        # Freeze panes + merge title
        ws.freeze_panes = ws.cell(header_row + 1, 1)
        if columns:
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1,   end_column=len(columns))
    else:
        # ── Plain output ───────────────────────────────────────────────────────
        ws.append(columns)
        for row in rows:
            ws.append([_safe_val(row, c) for c in columns])

    # Auto-width (always)
    data_rows = rows  # reference for width calculation
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in data_rows:
            val = row.get(col_name)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


# ── Auth routes ────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        users    = load_users()

        if username in users and check_password_hash(users[username], password):
            session["username"] = username
            audit("LOGIN", "success")
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        else:
            audit("LOGIN_FAIL", f"username={username}")
            error = "Неверный логин или пароль"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    audit("LOGOUT")
    session.clear()
    return redirect(url_for("login"))


# ── Main routes ────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return render_template("index.html", scripts=load_scripts(),
                           username=session.get("username"))


@app.route("/api/scripts")
@login_required
def api_scripts():
    return jsonify(load_scripts())


@app.route("/api/generate", methods=["POST"])
@login_required
def api_generate():
    data       = request.json or {}
    script_ids = data.get("scripts", [])
    date_from  = data.get("date_from", "")
    date_to    = data.get("date_to", "")
    styled     = bool(data.get("styled", False))
    use_time   = bool(data.get("use_time", True))
    time_from  = data.get("time_from", "00:00:00")
    time_to    = data.get("time_to",   "23:59:59")

    if not script_ids:
        return jsonify({"error": "Не выбрано ни одного скрипта"}), 400
    if not date_from or not date_to:
        return jsonify({"error": "Укажите диапазон дат"}), 400

    try:
        d_from = datetime.strptime(date_from, "%Y-%m-%d")
        d_to   = datetime.strptime(date_to,   "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Неверный формат дат"}), 400

    if d_to < d_from:
        return jsonify({"error": "Дата окончания раньше даты начала"}), 400
    if (d_to - d_from).days + 1 > 90:
        return jsonify({"error": "Диапазон дат не должен превышать 90 дней"}), 400

    all_scripts = {s["id"]: s for s in load_scripts()}
    selected    = [all_scripts[sid] for sid in script_ids if sid in all_scripts]
    if not selected:
        return jsonify({"error": "Выбранные скрипты не найдены"}), 400

    results     = {}
    empty_names = []
    errors      = []

    for script in selected:
        sid = script["id"]
        try:
            cols, rows = run_sql_script(script["file"], date_from, date_to,
                                               use_time=use_time,
                                               time_from=time_from, time_to=time_to)
            if rows:
                results[sid] = (script["name"], cols, rows)
            else:
                empty_names.append(script["name"])
        except Exception as exc:
            errors.append(f"{script['name']}: {exc}")

    # ── Audit log ──────────────────────────────────────────────────────────────
    requested_names = [all_scripts[sid]["name"] for sid in script_ids if sid in all_scripts]
    filled_names    = [name for name, _, _ in results.values()]
    audit(
        "GENERATE",
        f"period={date_from}/{date_to} "
        f"requested=[{', '.join(requested_names)}] "
        f"with_data=[{', '.join(filled_names)}] "
        f"empty=[{', '.join(empty_names)}]"
    )

    if errors:
        return jsonify({"error": "Ошибки выполнения скриптов:\n" + "\n".join(errors)}), 500

    if not results:
        return jsonify({
            "empty":   empty_names,
            "message": "Все выбранные скрипты не вернули данных за указанный период."
        }), 200

    label = f"{date_from}__{date_to}"

    if len(results) == 1:
        sid, (name, cols, rows) = next(iter(results.items()))
        wb = Workbook()
        ws = wb.active
        ws.title = name[:31]
        write_sheet(ws, cols, rows, styled=styled, title=name, date_from=date_from, date_to=date_to)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        filename = f"{sid}_{label}.xlsx"
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename,
        ), 200, {"X-Empty-Scripts": base64.b64encode(json.dumps(empty_names, ensure_ascii=False).encode()).decode()}

    else:
        wb = Workbook(); wb.remove(wb.active)
        for sid, (name, cols, rows) in results.items():
            ws = wb.create_sheet(title=name[:31])
            write_sheet(ws, cols, rows, styled=styled, title=name, date_from=date_from, date_to=date_to)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        filename = f"reports_{label}.xlsx"
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename,
        ), 200, {"X-Empty-Scripts": base64.b64encode(json.dumps(empty_names, ensure_ascii=False).encode()).decode()}


if __name__ == "__main__":
    app.run(debug=True, port=5000)
